from __future__ import annotations

from io import BytesIO
from datetime import datetime
import os
import re
import time
import xml.etree.ElementTree as ET

import anthropic
from dotenv import load_dotenv
import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers, Font, Alignment
from openpyxl.utils import get_column_letter
from audit_validation import validate_audit_sentences_sheet

try:
    from openai import OpenAI
except ImportError:  # Optional dependency
    OpenAI = None


DEFAULT_MAX_CATEGORIES = 1000
DEFAULT_MAX_SENTENCES_PER_CATEGORY = 51
DEFAULT_ANTHROPIC_MODEL = "claude-opus-4-5"
DEFAULT_OPENAI_MODEL = "gpt-5-nano"
DEFAULT_MAX_TOKENS = 10000
COLUMN_WIDTH_PX = 300
COLUMN_WIDTH_CHAR = round(COLUMN_WIDTH_PX / 7.0, 2)
FINDINGS_HEADERS = ["Topic", "Description", "Accuracy", "Issues"]
SENTENCES_HEADERS = ["ID", "Sentence", "Topic", "Audit", "Explanation"]
SETTINGS_HEADERS = ["Setting", "Value"]
ERRORS_HEADERS = ["Type", "Message"]
FINDINGS_WRAP_COLUMNS = (1, 2, 4)
SENTENCES_WRAP_COLUMNS = (2, 3, 5)
SETTINGS_WRAP_COLUMNS = (1, 2)
ERRORS_WRAP_COLUMNS = (1, 2)
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _apply_precision_formula(ws_categories, row_idx, sentences_sheet_title):
    sheet_ref = f"'{sentences_sheet_title}'"
    # Sentences columns: A=ID, B=Sentence, C=Topic, D=Audit, E=Explanation
    category_col = f"{sheet_ref}!C:C"
    judgment_col = f"{sheet_ref}!D:D"
    category_cell = "INDEX(A:A, ROW())"
    formula = (
        f"=COUNTIFS({category_col}, {category_cell}, {judgment_col}, \"YES\")"
        f"/COUNTIFS({category_col}, {category_cell}, {judgment_col}, \"<>\")"
    )
    cell = ws_categories.cell(row=row_idx, column=3)
    cell.value = formula
    cell.number_format = numbers.FORMAT_PERCENTAGE


def _add_model_average_row(ws_categories):
    ws_categories.insert_rows(2)
    ws_categories.cell(row=2, column=1, value="AVERAGE")
    last_row = ws_categories.max_row
    average_cell = ws_categories.cell(row=2, column=3)
    average_cell.value = f"=AVERAGE(C3:C{last_row})"
    average_cell.number_format = numbers.FORMAT_PERCENTAGE


def _ensure_headers(ws, headers):
    is_blank_header = (
        ws.max_row <= 1
        and all(
            (ws.cell(row=1, column=idx).value in (None, ""))
            for idx in range(1, len(headers) + 1)
        )
    )
    if is_blank_header:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
    else:
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)


def _apply_header_style(ws, header_len):
    for col_idx in range(1, header_len + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


def _apply_alignment_to_row(ws, row_idx, columns):
    for col_idx in columns:
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = WRAP_ALIGNMENT


def _apply_alignment_to_columns(ws, columns):
    if ws.max_row < 1:
        return
    for row_idx in range(1, ws.max_row + 1):
        _apply_alignment_to_row(ws, row_idx, columns)


def _set_column_widths(ws, columns):
    for col_idx in columns:
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTH_CHAR


def _refresh_auto_filter(ws):
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _ensure_findings_sheet(wb):
    if "Findings" in wb.sheetnames:
        ws = wb["Findings"]
    elif "Topics" in wb.sheetnames:
        ws = wb["Topics"]
        ws.title = "Findings"
    else:
        if wb.sheetnames:
            default_ws = wb[wb.sheetnames[0]]
            if default_ws.title == "Sentences":
                ws = wb.create_sheet(title="Findings", index=0)
            else:
                default_ws.title = "Findings"
                ws = default_ws
        else:
            ws = wb.create_sheet(title="Findings", index=0)

    _ensure_headers(ws, FINDINGS_HEADERS)
    _apply_header_style(ws, len(FINDINGS_HEADERS))
    _set_column_widths(ws, FINDINGS_WRAP_COLUMNS)
    _apply_alignment_to_columns(ws, FINDINGS_WRAP_COLUMNS)
    ws.freeze_panes = "A2"
    current_index = wb.sheetnames.index(ws.title)
    if current_index != 0:
        wb.move_sheet(ws, -current_index)
    return ws


def _ensure_sentences_sheet(wb):
    if "Sentences" in wb.sheetnames:
        ws = wb["Sentences"]
    else:
        ws = wb.create_sheet(title="Sentences")

    _ensure_headers(ws, SENTENCES_HEADERS)
    _apply_header_style(ws, len(SENTENCES_HEADERS))
    _set_column_widths(ws, SENTENCES_WRAP_COLUMNS)
    _apply_alignment_to_columns(ws, SENTENCES_WRAP_COLUMNS)
    ws.freeze_panes = "A2"
    return ws


def _format_categories_selected(topics_to_audit, selected_count, total_count):
    """Format the categories selected for display in the settings sheet."""
    if not topics_to_audit:
        return f"All ({total_count})"
    return f"{selected_count} of {total_count}"


def _ensure_settings_sheet(wb):
    if "Audit Settings" in wb.sheetnames:
        ws = wb["Audit Settings"]
    else:
        ws = wb.create_sheet(title="Audit Settings")

    _ensure_headers(ws, SETTINGS_HEADERS)
    _apply_header_style(ws, len(SETTINGS_HEADERS))
    _set_column_widths(ws, SETTINGS_WRAP_COLUMNS)
    _apply_alignment_to_columns(ws, SETTINGS_WRAP_COLUMNS)
    ws.freeze_panes = "A2"
    return ws


def _write_errors_sheet(wb, collected_warnings):
    """Create and populate the Errors sheet only if there are warnings to record."""
    if not collected_warnings:
        # Remove the sheet if it exists but there's nothing to write
        if "Errors" in wb.sheetnames:
            del wb["Errors"]
        return
    if "Errors" in wb.sheetnames:
        ws = wb["Errors"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet(title="Errors")
    _ensure_headers(ws, ERRORS_HEADERS)
    _apply_header_style(ws, len(ERRORS_HEADERS))
    _set_column_widths(ws, ERRORS_WRAP_COLUMNS)
    ws.freeze_panes = "A2"
    for msg in collected_warnings:
        ws.append(["Warning", msg])
        _apply_alignment_to_row(ws, ws.max_row, ERRORS_WRAP_COLUMNS)


def _write_settings_sheet(ws, settings):
    """Write key-value settings to the Audit Settings sheet, replacing any existing data."""
    # Clear existing data rows (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for key, value in settings.items():
        ws.append([key, value])
        _apply_alignment_to_row(ws, ws.max_row, SETTINGS_WRAP_COLUMNS)


def _update_setting(ws, key, value):
    """Update a single setting value by key in the Audit Settings sheet."""
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == key:
            ws.cell(row=row_idx, column=2, value=value)
            return


def _load_yaml(path):
    with open(path, 'r') as f:
        return yaml.safe_load(f) or {}


def _get_llm_client(llm_provider, anthropic_api_key=None, openai_api_key=None):
    load_dotenv()

    if llm_provider == 'anthropic':
        api_key = anthropic_api_key or os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise RuntimeError("ANTHROPIC_API_KEY environment variable not set")
        return anthropic.Anthropic(api_key=api_key)

    if llm_provider == 'openai':
        if OpenAI is None:
            raise RuntimeError("openai package is not installed")
        api_key = openai_api_key or os.getenv('OPENAI_API_KEY')
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY environment variable not set")
        return OpenAI(api_key=api_key)

    raise ValueError("llm_provider must be 'anthropic' or 'openai'")


def _build_category_sentences(df, is_output_format=False):
    """
    Build a mapping of category -> {sentence_id: sentence}.

    For input format: columns are [ID, Sentence, Category, ...]
    For output format (new): columns are [ID, Sentence, Topic, Audit, Explanation]
    For output format (old): columns are [Sentence, Topic, Audit, Explanation]
        - For old format, IDs are generated as row indices
    """
    category_sentences = {}

    if is_output_format:
        # Check if this is new format (with ID) or old format (without)
        first_col_name = str(df.columns[0]).strip().casefold() if len(df.columns) > 0 else ""
        has_id_col = first_col_name == "id"

        if has_id_col:
            id_col = df.columns[0]
            sentence_col = df.columns[1]
            category_col = df.columns[2]

            for _, row in df.iterrows():
                category = row[category_col]
                sentence_id = row[id_col]
                sentence = row[sentence_col]
                if pd.isna(category) or pd.isna(sentence) or pd.isna(sentence_id):
                    continue
                if category not in category_sentences:
                    category_sentences[category] = {}
                category_sentences[category][sentence_id] = sentence
        else:
            # Old output format without ID column
            sentence_col = df.columns[0]
            category_col = df.columns[1]
            # Generate synthetic IDs for old output format
            row_idx = 0
            for _, row in df.iterrows():
                row_idx += 1
                category = row[category_col]
                sentence = row[sentence_col]
                if pd.isna(category) or pd.isna(sentence):
                    continue
                if category not in category_sentences:
                    category_sentences[category] = {}
                # Use row index as synthetic ID
                category_sentences[category][row_idx] = sentence
    else:
        id_col = df.columns[0]
        sentence_col = df.columns[1]
        category_col = df.columns[2]

        first_row = df.iloc[0]
        if all(pd.isna(val) for val in first_row):
            raise ValueError(
                "The first row in the audit file is completely blank. "
                "Reformat the audit file before running the audit."
            )

        for _, row in df.iterrows():
            category = row[category_col]
            sentence_id = row[id_col]
            sentence = row[sentence_col]
            if pd.isna(category) or pd.isna(sentence) or pd.isna(sentence_id):
                continue
            if category not in category_sentences:
                category_sentences[category] = {}
            category_sentences[category][sentence_id] = sentence

    return category_sentences


def _extract_category_descriptions(model_tree_bytes):
    if not model_tree_bytes:
        return {}

    category_descriptions = {}
    root = ET.fromstring(model_tree_bytes)
    model_root = root if root.tag == "model" else root.find("model")
    tree_elem = None
    if model_root is not None:
        tree_elem = model_root.find('tree')
    if tree_elem is None:
        tree_elem = root.find('tree')

    root_node = tree_elem[0] if tree_elem is not None and len(tree_elem) > 0 else None
    if root_node is None:
        return category_descriptions

    root_name = root_node.get('name')

    def get_all_child_names_and_descriptions(parent, path_so_far=None):
        if path_so_far is None:
            path_so_far = []
        cat_name = parent.get('name')
        if cat_name is None:
            return
        if len(path_so_far) > 0 and root_name and path_so_far[0] == root_name:
            path_so_far = path_so_far[1:]
        updated_path = path_so_far + [cat_name]
        full_cat_path = "-->".join(updated_path)
        category_desc = parent.get('description') or "None"
        category_descriptions[full_cat_path] = category_desc
        for child in parent:
            get_all_child_names_and_descriptions(child, updated_path)

    get_all_child_names_and_descriptions(root_node)
    return category_descriptions


def _parse_llm_response(response_text):
    pattern = r"ID:\s*(.+?)\s*-\s*Judgment:\s*(YES|NO)\s*-\s*Reasoning:\s*(.+)"
    matches = re.findall(pattern, response_text, re.IGNORECASE)
    nlp_results = {}
    for match in matches:
        sent_id = str(match[0]).strip()
        judgment = match[1].strip().upper()
        explanation = match[2].strip()
        nlp_results[sent_id] = (judgment, explanation)
    return nlp_results


class AuditStopRequested(Exception):
    """Raised when the user requests to stop the audit."""
    pass


def detect_partial_audit(audit_bytes):
    """
    Detect if the uploaded file is a partial audit output from this app.

    Returns a dict with:
      - is_partial: bool - True if this is a partial audit
      - completed_categories: set - categories with all sentences judged
      - incomplete_categories: set - categories with some but not all sentences judged
      - unjudged_categories: set - categories with no sentences judged
    """
    result = {
        "is_partial": False,
        "completed_categories": set(),
        "incomplete_categories": set(),
        "unjudged_categories": set(),
    }

    try:
        excel_file = pd.ExcelFile(BytesIO(audit_bytes))
        sheet_names_lower = [name.casefold() for name in excel_file.sheet_names]

        # Check if this looks like our output format (has Findings and Sentences sheets)
        has_findings = "findings" in sheet_names_lower or "topics" in sheet_names_lower
        has_sentences = "sentences" in sheet_names_lower

        if not (has_findings and has_sentences):
            return result

        # Find the actual sheet names
        sentences_sheet = None
        for name in excel_file.sheet_names:
            if name.casefold() == "sentences":
                sentences_sheet = name
                break

        if not sentences_sheet:
            return result

        df = pd.read_excel(excel_file, sheet_name=sentences_sheet)

        # Check if it has expected headers: ID, Sentence, Topic, Audit, Explanation
        # Also support old format without ID: Sentence, Topic, Audit, Explanation
        if len(df.columns) < 4:
            return result

        headers = [str(col).strip().casefold() for col in df.columns[:5] if col is not None]
        expected_headers_new = ["id", "sentence", "topic", "audit", "explanation"]
        expected_headers_old = ["sentence", "topic", "audit", "explanation"]

        has_id_column = False
        if len(headers) >= 5 and headers[:5] == expected_headers_new:
            has_id_column = True
        elif len(headers) >= 4 and headers[:4] == expected_headers_old:
            has_id_column = False
        else:
            return result

        # This is our output format - analyze category completion
        result["is_partial"] = True

        if has_id_column:
            id_col = df.columns[0]
            sentence_col = df.columns[1]
            topic_col = df.columns[2]
            audit_col = df.columns[3]
            explanation_col = df.columns[4]
        else:
            id_col = None
            sentence_col = df.columns[0]
            topic_col = df.columns[1]
            audit_col = df.columns[2]
            explanation_col = df.columns[3]

        # Group sentences by category and check completion
        category_stats = {}
        for _, row in df.iterrows():
            topic = row[topic_col]
            if pd.isna(topic):
                continue
            topic_str = str(topic).strip()
            if not topic_str:
                continue

            if topic_str not in category_stats:
                category_stats[topic_str] = {"total": 0, "judged": 0}

            category_stats[topic_str]["total"] += 1

            # A sentence is considered judged if it has a non-empty Explanation
            explanation = row[explanation_col]
            if not pd.isna(explanation) and str(explanation).strip():
                category_stats[topic_str]["judged"] += 1

        for category, stats in category_stats.items():
            if stats["judged"] == 0:
                result["unjudged_categories"].add(category)
            elif stats["judged"] < stats["total"]:
                result["incomplete_categories"].add(category)
            else:
                result["completed_categories"].add(category)

        # If all categories are complete, this is not a partial audit
        if not result["incomplete_categories"] and not result["unjudged_categories"]:
            result["is_partial"] = False

    except Exception:
        pass

    return result


def _load_existing_audit_data(audit_bytes):
    """
    Load existing audit data from a partial audit file.

    Returns a dict with:
      - sentences_by_category: {category: [(sentence_id, sentence, judgment, explanation), ...]}
      - findings_by_category: {category: description}
    """
    result = {
        "sentences_by_category": {},
        "findings_by_category": {},
    }

    try:
        wb = load_workbook(BytesIO(audit_bytes))

        # Load sentences - detect if ID column exists
        if "Sentences" in wb.sheetnames:
            ws = wb["Sentences"]
            # Check header to determine format
            first_header = ws.cell(row=1, column=1).value
            has_id_col = first_header and str(first_header).strip().casefold() == "id"

            for row_idx in range(2, ws.max_row + 1):
                if has_id_col:
                    sentence_id = ws.cell(row=row_idx, column=1).value
                    sentence = ws.cell(row=row_idx, column=2).value
                    topic = ws.cell(row=row_idx, column=3).value
                    judgment = ws.cell(row=row_idx, column=4).value
                    explanation = ws.cell(row=row_idx, column=5).value
                else:
                    # Old format without ID - use row index as synthetic ID
                    sentence_id = row_idx - 1
                    sentence = ws.cell(row=row_idx, column=1).value
                    topic = ws.cell(row=row_idx, column=2).value
                    judgment = ws.cell(row=row_idx, column=3).value
                    explanation = ws.cell(row=row_idx, column=4).value

                if topic is None:
                    continue
                topic_str = str(topic).strip()
                if not topic_str:
                    continue

                if topic_str not in result["sentences_by_category"]:
                    result["sentences_by_category"][topic_str] = []

                result["sentences_by_category"][topic_str].append((
                    sentence_id,
                    sentence,
                    judgment if judgment else "",
                    explanation if explanation else "",
                ))

        # Load findings (skip AVERAGE row)
        findings_sheet = None
        for name in ["Findings", "Topics"]:
            if name in wb.sheetnames:
                findings_sheet = name
                break

        if findings_sheet:
            ws = wb[findings_sheet]
            for row_idx in range(2, ws.max_row + 1):
                topic = ws.cell(row=row_idx, column=1).value
                if topic is None:
                    continue
                topic_str = str(topic).strip()
                if topic_str.upper() == "AVERAGE":
                    continue
                description = ws.cell(row=row_idx, column=2).value
                result["findings_by_category"][topic_str] = description if description else ""

        wb.close()
    except Exception:
        pass

    return result


def _is_retryable_llm_error(exc):
    """Return True if the exception is a transient LLM API error worth retrying."""
    status = getattr(exc, "status_code", None) or getattr(exc, "status", None)
    if status in (429, 529, 503):
        return True
    msg = str(exc).lower()
    if any(keyword in msg for keyword in ("overloaded", "rate limit", "too many requests", "service unavailable")):
        return True
    return False


def run_audit(
    audit_excel_bytes,
    prompt_template,
    llm_provider,
    model_name=None,
    model_info="",
    organization="the organization",
    audience="feedback",
    max_categories=DEFAULT_MAX_CATEGORIES,
    max_sentences_per_category=DEFAULT_MAX_SENTENCES_PER_CATEGORY,
    model_tree_bytes=None,
    topics_to_audit=None,
    anthropic_api_key=None,
    openai_api_key=None,
    max_tokens=DEFAULT_MAX_TOKENS,
    log_fn=None,
    warn_fn=None,
    progress_fn=None,
    save_progress_fn=None,
    check_stop_fn=None,
    existing_audit_bytes=None,
    completed_categories=None,
    audit_file_name=None,
    model_tree_name=None,
    include_summary=False,
    summary_prompt="",
    accuracy_threshold=0.80,
    run_datetime=None,
    audit_warnings=None,
):
    if log_fn is None:
        log_fn = lambda *_args, **_kwargs: None
    if warn_fn is None:
        warn_fn = log_fn

    sentences_sheet, header_row_idx, _, warnings, is_output_format = validate_audit_sentences_sheet(audit_excel_bytes)
    if warnings:
        warn_fn("Input audit file warnings:\n" + "\n".join(warnings))
    df = pd.read_excel(
        BytesIO(audit_excel_bytes),
        sheet_name=sentences_sheet,
        header=header_row_idx,
    )
    category_sentences = _build_category_sentences(df, is_output_format)

    category_descriptions = _extract_category_descriptions(model_tree_bytes)

    all_categories = list(category_sentences.keys())
    total_category_count = len(all_categories)
    categories_to_audit = all_categories
    if topics_to_audit:
        def _normalize_topic(value):
            text = str(value).strip()
            return " ".join(text.split())

        def _key(value):
            text = _normalize_topic(value)
            parts = [part.strip() for part in re.split(r"\s*-->\s*", text) if part.strip()]
            normalized = "-->".join(parts) if parts else text
            return normalized.casefold()

        audit_categories_by_key = { _key(cat): cat for cat in categories_to_audit }

        filtered_categories = []
        seen = set()
        unmatched_topics = []
        for topic in topics_to_audit:
            topic_key = _key(topic)
            match = None
            if topic_key in audit_categories_by_key:
                match = [audit_categories_by_key[topic_key]]

            if match:
                for cat in match:
                    if cat not in seen:
                        filtered_categories.append(cat)
                        seen.add(cat)
            else:
                unmatched_topics.append(topic)

        # if unmatched_topics:
        #     unmatched_list = "\n".join(f"- {topic}" for topic in unmatched_topics)
        #     warn_fn(
        #         "Warning: Sentences for some categories not found in audit file.\n"
        #         f"{unmatched_list}"
        #     )

        if filtered_categories:
            categories_to_audit = filtered_categories
        else:
            raise ValueError(
                "No selected topics matched the audit categories. "
                "Check the selection or upload a matching model tree."
            )

    if model_name is None:
        model_name = DEFAULT_ANTHROPIC_MODEL if llm_provider == 'anthropic' else DEFAULT_OPENAI_MODEL

    # Prepare sets for tracking which categories to actually audit
    if completed_categories is None:
        completed_categories = set()

    # Load existing audit data if resuming
    existing_data = None
    if existing_audit_bytes:
        existing_data = _load_existing_audit_data(existing_audit_bytes)

    wb = Workbook()
    ws_findings = _ensure_findings_sheet(wb)
    ws_sentences = _ensure_sentences_sheet(wb)
    ws_settings = _ensure_settings_sheet(wb)

    settings = {
        "Input File": audit_file_name or "",
        "LLM Provider": llm_provider,
        "Model": model_name,
        "Organization": organization,
        "Audience": audience,
        "Context": model_info or "",
        "Model Tree File": model_tree_name or "(none)",
        "Categories Selected": _format_categories_selected(topics_to_audit, len(categories_to_audit), total_category_count),
        "Audit Prompt": prompt_template,
        "Include Summary of Issues": "Yes" if include_summary else "No",
        "Summary Prompt": summary_prompt if include_summary else "(n/a)",
        "Accuracy Threshold": accuracy_threshold if include_summary else "(n/a)",
        "Max Categories": max_categories,
        "Max Sentences per Category": max_sentences_per_category,
        "Max Tokens per Request": max_tokens,
        "Run Started": run_datetime or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Run Finished": "",
    }
    _write_settings_sheet(ws_settings, settings)

    # Collect all warnings (pre-flight + runtime) for the Errors sheet
    collected_warnings = list(audit_warnings or [])

    _original_warn_fn = warn_fn
    def _collecting_warn_fn(msg):
        collected_warnings.append(str(msg))
        _original_warn_fn(msg)
    warn_fn = _collecting_warn_fn

    # Track which categories need LLM auditing vs already completed
    categories_needing_audit = []
    for category in categories_to_audit:
        if category in completed_categories:
            continue
        categories_needing_audit.append(category)

    # Write all categories upfront (completed ones with their data, others with placeholders)
    findings_row_map = {}  # category -> row index in findings sheet
    sentences_row_ranges = {}  # category -> (start_row, end_row) in sentences sheet

    for category in categories_to_audit:
        description = category_descriptions.get(category, "None") or "None"
        sent_tuples = list(category_sentences[category].items())

        # Record starting row for sentences
        start_row = ws_sentences.max_row + 1

        if category in completed_categories and existing_data:
            # Use existing data for completed categories
            existing_sentences = existing_data["sentences_by_category"].get(category, [])
            for sent_data in existing_sentences:
                sentence_id, sentence, judgment, explanation = sent_data
                ws_sentences.append([sentence_id, sentence, category, judgment, explanation])
                _apply_alignment_to_row(ws_sentences, ws_sentences.max_row, SENTENCES_WRAP_COLUMNS)

            # Add findings row
            ws_findings.append([category, description, "", ""])
            _apply_alignment_to_row(ws_findings, ws_findings.max_row, FINDINGS_WRAP_COLUMNS)
            _apply_precision_formula(ws_findings, ws_findings.max_row, ws_sentences.title)
        else:
            # Write placeholder sentences for categories to be audited
            for sentence_id, sentence in sent_tuples:
                ws_sentences.append([sentence_id, sentence, category, "", ""])
                _apply_alignment_to_row(ws_sentences, ws_sentences.max_row, SENTENCES_WRAP_COLUMNS)

            # Add findings row with "Not yet audited" status
            ws_findings.append([category, description, "", "Not yet audited"])
            _apply_alignment_to_row(ws_findings, ws_findings.max_row, FINDINGS_WRAP_COLUMNS)
            _apply_precision_formula(ws_findings, ws_findings.max_row, ws_sentences.title)

        end_row = ws_sentences.max_row
        findings_row_map[category] = ws_findings.max_row
        sentences_row_ranges[category] = (start_row, end_row)

    def _save_current_workbook():
        """Save current workbook state and return bytes."""
        added_model_avg = False
        if ws_findings.max_row > 1:
            # Temporarily add average row for the partial output
            _add_model_average_row(ws_findings)
            _apply_alignment_to_row(ws_findings, 2, FINDINGS_WRAP_COLUMNS)
            added_model_avg = True
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _update_setting(ws_settings, "Run Finished", f"In progress (last saved: {now})")
        _write_errors_sheet(wb, collected_warnings)
        _refresh_auto_filter(ws_findings)
        _refresh_auto_filter(ws_sentences)
        temp_output = BytesIO()
        wb.save(temp_output)
        temp_output.seek(0)
        result = temp_output.getvalue()
        # Remove the temporarily added average row
        if added_model_avg:
            ws_findings.delete_rows(2)
        return result

    # Now only connect to LLM if there are categories to audit
    client = None
    if categories_needing_audit:
        client = _get_llm_client(llm_provider, anthropic_api_key, openai_api_key)

    total_categories = min(len(categories_needing_audit), max_categories)
    cat_count = 0
    for category in categories_needing_audit:
        cat_count += 1
        if cat_count > max_categories:
            log_fn("Reached max_categories limit.")
            break

        # Check if stop was requested
        if check_stop_fn and check_stop_fn():
            if save_progress_fn:
                save_progress_fn(_save_current_workbook())
            raise AuditStopRequested("Audit stopped by user request")

        if progress_fn:
            progress_fn(cat_count, total_categories, category)

        description = category_descriptions.get(category, "None") or "None"

        sent_tuples = list(category_sentences[category].items())
        sentences_text = ""
        sent_count = 0
        for sentence_id, sentence in sent_tuples:
            sent_count += 1
            if sent_count > max_sentences_per_category:
                break
            sentences_text += f"ID: {sentence_id} - {sentence}\n"

        message_content = prompt_template.format(
            category=category,
            description=description,
            sentences_text=sentences_text,
            model_info=model_info or "",
            organization=organization,
            audience=audience,
        )

        # log_fn(f"Sending message to LLM for category {category}...")

        retry_delays = [30, 60, 120]
        response_text = None
        for attempt in range(len(retry_delays) + 1):
            try:
                if llm_provider == 'anthropic':
                    message = client.messages.create(
                        model=model_name,
                        max_tokens=max_tokens,
                        messages=[
                            {"role": "user", "content": message_content}
                        ]
                    )
                    response_text = message.content[0].text
                elif llm_provider == 'openai':
                    response = client.chat.completions.create(
                        model=model_name,
                        max_completion_tokens=max_tokens,
                        messages=[
                            {"role": "user", "content": message_content}
                        ]
                    )
                    response_text = response.choices[0].message.content
                else:
                    raise ValueError("llm_provider must be 'anthropic' or 'openai'")
                break  # Success, exit retry loop
            except Exception as e:
                is_retryable = _is_retryable_llm_error(e)
                if is_retryable and attempt < len(retry_delays):
                    delay = retry_delays[attempt]
                    if log_fn:
                        log_fn(
                            f"LLM API returned a retryable error (attempt {attempt + 1}/{len(retry_delays) + 1}). "
                            f"Retrying in {delay} seconds..."
                        )
                    if save_progress_fn:
                        save_progress_fn(_save_current_workbook())
                    time.sleep(delay)
                    continue
                # Non-retryable error or exhausted retries — save progress and re-raise
                if save_progress_fn:
                    save_progress_fn(_save_current_workbook())
                raise

        nlp_results = _parse_llm_response(response_text)

        # Validate LLM response
        sent_ids = {str(sid) for sid, _ in sent_tuples[:max_sentences_per_category]}
        returned_ids = set(nlp_results.keys())
        if not returned_ids:
            warn_fn(f"Category \"{category}\": LLM response could not be parsed. All sentences will have blank audit results.")
        else:
            missing = sent_ids - returned_ids
            extra = returned_ids - sent_ids
            if missing:
                warn_fn(f"Category \"{category}\": {len(missing)} of {len(sent_ids)} sentences missing from LLM response.")
            if extra:
                warn_fn(f"Category \"{category}\": LLM returned {len(extra)} unrecognized sentence IDs.")

        # Update the existing sentences rows with judgment results
        # Columns: ID(1), Sentence(2), Topic(3), Audit(4), Explanation(5)
        start_row, end_row = sentences_row_ranges[category]
        row_idx = start_row
        for sentence_id, sentence in sent_tuples:
            if row_idx > end_row:
                break
            judgment, explanation = nlp_results.get(str(sentence_id), ("", ""))
            ws_sentences.cell(row=row_idx, column=4, value=judgment)
            ws_sentences.cell(row=row_idx, column=5, value=explanation)
            _apply_alignment_to_row(ws_sentences, row_idx, SENTENCES_WRAP_COLUMNS)
            row_idx += 1

        # Update the findings row to clear "Not yet audited" status
        findings_row = findings_row_map[category]
        ws_findings.cell(row=findings_row, column=4, value="")

        # Save progress after each category completes
        if save_progress_fn:
            save_progress_fn(_save_current_workbook())

    # Record finish time
    _update_setting(ws_settings, "Run Finished", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    _write_errors_sheet(wb, collected_warnings)

    if ws_findings.max_row > 1:
        _add_model_average_row(ws_findings)
        _apply_alignment_to_row(ws_findings, 2, FINDINGS_WRAP_COLUMNS)

    _refresh_auto_filter(ws_findings)
    _refresh_auto_filter(ws_sentences)

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.getvalue()


def run_audit_from_config():
    print("Starting script...")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    prompts_path = os.path.join(script_dir, 'prompts.yaml')
    config_path = os.path.join(script_dir, 'config.yaml')
    inputs_dir = os.path.join(script_dir, "inputs")

    try:
        prompts = _load_yaml(prompts_path)
        config = _load_yaml(config_path)
        audit_config = config.get('cli_audit', {})
        max_categories = audit_config.get('max_categories', DEFAULT_MAX_CATEGORIES)
        max_sentences_per_category = audit_config.get(
            'max_sentences_per_category',
            DEFAULT_MAX_SENTENCES_PER_CATEGORY,
        )
        max_tokens = audit_config.get('max_tokens', DEFAULT_MAX_TOKENS)
        organization = audit_config.get('organization', 'the organization')
        audience = audit_config.get('audience', 'feedback')
        msg_template = prompts.get('audit_prompt', '')
        audit_file_name = audit_config.get('audit_file')
        model_tree_file = audit_config.get('model_tree')
        audit_in_progress = audit_config.get('audit_in_progress')
        llm_provider = audit_config.get('llm_provider', 'anthropic')
        model_name = audit_config.get('model_name')
    except FileNotFoundError:
        print("Error: prompts.yaml or config.yaml not found")
        return

    if llm_provider not in ["anthropic", "openai"]:
        print("Error: llm_provider not properly set in config file. Use 'anthropic' or 'openai'.")
        return
    if not audit_file_name:
        print("Error: cli_audit.audit_file is missing in config.yaml")
        return

    client = _get_llm_client(llm_provider)
    print("Retrieved API key")

    excel_path = os.path.join(inputs_dir, audit_file_name)
    with open(excel_path, "rb") as f:
        file_bytes = f.read()
    sentences_sheet, header_row_idx, _, warnings, is_output_format = validate_audit_sentences_sheet(file_bytes)
    if warnings:
        for warning in warnings:
            print(f"WARNING: {warning}")
    df = pd.read_excel(excel_path, sheet_name=sentences_sheet, header=header_row_idx)
    category_sentences = _build_category_sentences(df, is_output_format)

    model_tree_bytes = None
    if model_tree_file:
        model_tree_path = os.path.join(inputs_dir, model_tree_file)
        if os.path.exists(model_tree_path):
            with open(model_tree_path, "rb") as f:
                model_tree_bytes = f.read()

    category_descriptions = _extract_category_descriptions(model_tree_bytes)

    outputs_dir = os.path.join(os.path.dirname(__file__), "outputs")
    if not os.path.exists(outputs_dir):
        os.makedirs(outputs_dir)

    timestamp = datetime.now().strftime("%y%m%d%H%M")
    resume_mode = False
    completed_categories = set()
    restart_category = None

    if audit_in_progress:
        in_progress_path = os.path.join(outputs_dir, audit_in_progress)
        if os.path.exists(in_progress_path):
            print(f"Resuming audit from in-progress file: {audit_in_progress}")
            output_path = in_progress_path
            resume_mode = True
        else:
            print(f"Provided audit_in_progress file '{audit_in_progress}' not found. Starting a new audit file.")

    if not resume_mode:
        output_filename = f"completed_audit_{timestamp}.xlsx"
        output_path = os.path.join(outputs_dir, output_filename)

    if resume_mode:
        wb = load_workbook(output_path)
    else:
        wb = Workbook()

    ws_findings = _ensure_findings_sheet(wb)
    ws_sentences = _ensure_sentences_sheet(wb)

    if resume_mode:
        existing_categories = [
            row[0]
            for row in ws_findings.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
            if row[0]
        ]
        if existing_categories:
            restart_category = existing_categories[-1]
            completed_categories = set(existing_categories[:-1])
            print(f"Last completed category recorded as '{restart_category}'. Re-auditing it before continuing.")

        if restart_category:
            for row_idx in range(ws_sentences.max_row, 1, -1):
                if ws_sentences.cell(row=row_idx, column=2).value == restart_category:
                    ws_sentences.delete_rows(row_idx)
            for row_idx in range(ws_findings.max_row, 1, -1):
                if ws_findings.cell(row=row_idx, column=1).value == restart_category:
                    ws_findings.delete_rows(row_idx)

    categories_to_audit = []
    if restart_category:
        categories_to_audit.append(restart_category)
    for category in category_sentences:
        if resume_mode and category in completed_categories:
            continue
        if restart_category and category == restart_category:
            continue
        categories_to_audit.append(category)

    cat_count = 0
    for category in categories_to_audit:
        cat_count += 1
        print(f"Auditing category {cat_count}, {category}")
        if cat_count >= max_categories:
            print("Reached max iteration")
            break

        description = category_descriptions.get(category, "None") or "None"

        sent_tuples = list(category_sentences[category].items())
        sentences_text = ""
        sent_count = 0
        for sentence_id, sentence in sent_tuples:
            sent_count += 1
            if sent_count > max_sentences_per_category:
                break
            sentences_text += f"ID: {sentence_id} - {sentence}\n"

        message_content = msg_template.format(
            category=category,
            description=description,
            sentences_text=sentences_text,
            organization=organization,
            audience=audience,
        )

        print(f"Sending message to LLM for category {category}...")

        if llm_provider == 'anthropic':
            message = client.messages.create(
                model=model_name or DEFAULT_ANTHROPIC_MODEL,
                max_tokens=max_tokens,
                messages=[
                    {"role": "user", "content": message_content}
                ]
            )
            response_text = message.content[0].text
        else:
            response = client.chat.completions.create(
                model=model_name or DEFAULT_OPENAI_MODEL,
                max_completion_tokens=max_tokens,
                messages=[
                    {"role": "user", "content": message_content}
                ]
            )
            response_text = response.choices[0].message.content

        print(f"Received response for {category}. Preview:")
        print(response_text[:50])

        nlp_results = _parse_llm_response(response_text)

        for sentence_id, sentence in sent_tuples:
            judgment, explanation = nlp_results.get(str(sentence_id), ("", ""))
            ws_sentences.append([sentence, category, judgment, explanation])
            _apply_alignment_to_row(ws_sentences, ws_sentences.max_row, SENTENCES_WRAP_COLUMNS)

        ws_findings.append([category, description, "", ""])
        _apply_alignment_to_row(ws_findings, ws_findings.max_row, FINDINGS_WRAP_COLUMNS)
        _apply_precision_formula(ws_findings, ws_findings.max_row, ws_sentences.title)
        _refresh_auto_filter(ws_findings)
        _refresh_auto_filter(ws_sentences)
        wb.save(output_path)

    if ws_findings.max_row > 1:
        _add_model_average_row(ws_findings)
        _apply_alignment_to_row(ws_findings, 2, FINDINGS_WRAP_COLUMNS)

    _refresh_auto_filter(ws_findings)
    _refresh_auto_filter(ws_sentences)
    wb.save(output_path)

    wb.close()
    print("Script concluded")


if __name__ == "__main__":
    run_audit_from_config()
