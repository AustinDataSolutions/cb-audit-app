"""Tests for audit.py — core audit engine."""
from __future__ import annotations

import threading
import time
from io import BytesIO
from unittest.mock import MagicMock, patch

import httpx
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers

from tests.conftest import _load_module, _make_workbook_bytes

audit = _load_module("audit.py", "audit")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_audit_workbook(sentences_data, findings_topics=None):
    """Build a minimal audit workbook from (id, sentence, topic, audit, explanation) tuples.

    If ``findings_topics`` is provided, it's an ordered list of topics that
    will be written to the Findings sheet column A. Otherwise the Findings
    sheet contains only a header row (legacy behavior).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sentences"
    ws.append(["ID", "Sentence", "Topic", "Audit", "Explanation"])
    for row in sentences_data:
        ws.append(list(row))

    ws_findings = wb.create_sheet("Findings")
    ws_findings.append(["Topic", "Description", "Accuracy", "Issues"])
    if findings_topics:
        for topic in findings_topics:
            ws_findings.append([topic, "None", "", "Not yet audited"])

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.getvalue()


# ===========================================================================
# _parse_llm_response
# ===========================================================================

class TestParseLlmResponse:
    def test_valid_response(self):
        response = (
            "ID: 1 - Judgment: YES - Reasoning: Looks accurate\n"
            "ID: 2 - Judgment: NO - Reasoning: Wrong category\n"
            "ID: 3 - Judgment: YES - Reasoning: Good fit\n"
        )
        result = audit._parse_llm_response(response)
        assert len(result) == 3
        assert result["1"] == ("YES", "Looks accurate")
        assert result["2"] == ("NO", "Wrong category")
        assert result["3"] == ("YES", "Good fit")

    def test_case_insensitive(self):
        response = (
            "ID: 1 - Judgment: yes - Reasoning: Fine\n"
            "ID: 2 - Judgment: No - Reasoning: Not good\n"
        )
        result = audit._parse_llm_response(response)
        assert result["1"] == ("YES", "Fine")
        assert result["2"] == ("NO", "Not good")

    def test_empty_response(self):
        assert audit._parse_llm_response("") == {}

    def test_malformed_lines(self):
        assert audit._parse_llm_response("This is not a valid response.") == {}

    def test_mixed_valid_and_invalid(self):
        response = (
            "ID: 1 - Judgment: YES - Reasoning: Good\n"
            "This line is garbage\n"
            "ID: 3 - Judgment: NO - Reasoning: Bad\n"
        )
        result = audit._parse_llm_response(response)
        assert len(result) == 2
        assert "1" in result
        assert "2" not in result
        assert "3" in result

    def test_unrecognized_ids(self):
        response = "ID: 999 - Judgment: YES - Reasoning: Phantom sentence\n"
        result = audit._parse_llm_response(response)
        assert result["999"] == ("YES", "Phantom sentence")

    def test_string_ids(self):
        response = "ID: abc-123 - Judgment: YES - Reasoning: Correct\n"
        result = audit._parse_llm_response(response)
        assert result["abc-123"] == ("YES", "Correct")

    def test_multiline_reasoning_stays_on_one_line(self):
        """Each line is parsed independently; reasoning stops at end of line."""
        response = "ID: 1 - Judgment: YES - Reasoning: Good match overall\n"
        result = audit._parse_llm_response(response)
        assert result["1"][1] == "Good match overall"


# ===========================================================================
# _is_retryable_llm_error
# ===========================================================================

class TestIsRetryableLlmError:
    def test_httpx_timeout(self):
        exc = httpx.TimeoutException("timed out")
        assert audit._is_retryable_llm_error(exc) is True

    def test_httpx_read_timeout(self):
        exc = httpx.ReadTimeout("read timed out")
        assert audit._is_retryable_llm_error(exc) is True

    def test_status_429(self):
        exc = Exception("Rate limited")
        exc.status_code = 429
        assert audit._is_retryable_llm_error(exc) is True

    def test_status_529(self):
        exc = Exception("Overloaded")
        exc.status_code = 529
        assert audit._is_retryable_llm_error(exc) is True

    def test_status_503(self):
        exc = Exception("Service unavailable")
        exc.status_code = 503
        assert audit._is_retryable_llm_error(exc) is True

    def test_status_attribute_via_status(self):
        """Some exceptions use .status instead of .status_code."""
        exc = Exception("err")
        exc.status = 429
        assert audit._is_retryable_llm_error(exc) is True

    def test_overloaded_keyword(self):
        exc = Exception("The API is overloaded right now")
        assert audit._is_retryable_llm_error(exc) is True

    def test_rate_limit_keyword(self):
        exc = Exception("rate limit exceeded")
        assert audit._is_retryable_llm_error(exc) is True

    def test_too_many_requests_keyword(self):
        exc = Exception("too many requests")
        assert audit._is_retryable_llm_error(exc) is True

    def test_service_unavailable_keyword(self):
        exc = Exception("service unavailable")
        assert audit._is_retryable_llm_error(exc) is True

    def test_timed_out_keyword(self):
        exc = Exception("request timed out")
        assert audit._is_retryable_llm_error(exc) is True

    def test_timeout_keyword(self):
        exc = Exception("timeout waiting for response")
        assert audit._is_retryable_llm_error(exc) is True

    def test_not_retryable_400(self):
        exc = Exception("Bad request")
        exc.status_code = 400
        assert audit._is_retryable_llm_error(exc) is False

    def test_not_retryable_401(self):
        exc = Exception("Unauthorized")
        exc.status_code = 401
        assert audit._is_retryable_llm_error(exc) is False

    def test_not_retryable_generic(self):
        exc = ValueError("Something went wrong")
        assert audit._is_retryable_llm_error(exc) is False


# ===========================================================================
# _call_llm_with_status
# ===========================================================================

class TestCallLlmWithStatus:
    def test_fast_call_no_status(self):
        """A call completing quickly should not trigger status messages."""
        status_calls = []
        result = audit._call_llm_with_status(
            lambda: "fast result", 300, lambda msg: status_calls.append(msg)
        )
        assert result == "fast result"
        # Only the clear call ("")
        assert status_calls == [""]

    def test_slow_call_triggers_status(self):
        """Monkeypatch the delay to a small value so we can test status messages."""
        original_delay = audit._LLM_STATUS_DELAY
        audit._LLM_STATUS_DELAY = 1  # 1 second

        status_calls = []

        def slow_fn():
            time.sleep(2.5)
            return "done"

        try:
            result = audit._call_llm_with_status(
                slow_fn, 300, lambda msg: status_calls.append(msg)
            )
            assert result == "done"
            # Should have at least one "Waiting..." message plus the clear
            waiting_msgs = [m for m in status_calls if m.startswith("Waiting")]
            assert len(waiting_msgs) >= 1
            assert status_calls[-1] == ""  # clear call
        finally:
            audit._LLM_STATUS_DELAY = original_delay

    def test_exception_is_reraised(self):
        def failing_fn():
            raise RuntimeError("LLM exploded")

        with pytest.raises(RuntimeError, match="LLM exploded"):
            audit._call_llm_with_status(failing_fn, 300, None)

    def test_exception_still_clears_status(self):
        status_calls = []

        def failing_fn():
            raise RuntimeError("boom")

        with pytest.raises(RuntimeError):
            audit._call_llm_with_status(
                failing_fn, 300, lambda msg: status_calls.append(msg)
            )
        # The clear call should still happen
        assert "" in status_calls

    def test_no_status_fn(self):
        """When status_fn is None, no crash occurs."""
        result = audit._call_llm_with_status(lambda: 42, 300, None)
        assert result == 42


# ===========================================================================
# detect_partial_audit
# ===========================================================================

class TestDetectPartialAudit:
    def test_fully_completed(self):
        wb_bytes = _build_audit_workbook([
            (1, "Sentence one", "TopicA", "YES", "Good"),
            (2, "Sentence two", "TopicA", "NO", "Bad"),
        ])
        result = audit.detect_partial_audit(wb_bytes)
        assert result["is_partial"] is False
        assert "TopicA" in result["completed_categories"]

    def test_partial_incomplete(self):
        wb_bytes = _build_audit_workbook([
            (1, "Sentence one", "TopicA", "YES", "Good"),
            (2, "Sentence two", "TopicA", "", ""),
            (3, "Sentence three", "TopicB", "", ""),
        ])
        result = audit.detect_partial_audit(wb_bytes)
        assert result["is_partial"] is True
        assert "TopicA" in result["incomplete_categories"]
        assert "TopicB" in result["unjudged_categories"]

    def test_not_audit_file(self):
        wb_bytes = _make_workbook_bytes({
            "Data": [["Col1", "Col2"], ["v1", "v2"]],
        })
        result = audit.detect_partial_audit(wb_bytes)
        assert result["is_partial"] is False

    def test_mixed_categories(self):
        wb_bytes = _build_audit_workbook([
            (1, "S1", "Complete", "YES", "Good"),
            (2, "S2", "Complete", "NO", "Bad"),
            (3, "S3", "Partial", "YES", "Fine"),
            (4, "S4", "Partial", "", ""),
            (5, "S5", "Unjudged", "", ""),
        ])
        result = audit.detect_partial_audit(wb_bytes)
        assert result["is_partial"] is True
        assert "Complete" in result["completed_categories"]
        assert "Partial" in result["incomplete_categories"]
        assert "Unjudged" in result["unjudged_categories"]

    def test_invalid_bytes(self):
        result = audit.detect_partial_audit(b"not an excel file")
        assert result["is_partial"] is False

    def test_selected_categories_from_findings(self):
        """selected_categories should be ordered by Findings col A, even when
        some categories have zero sentences (which can happen for topics
        with no rules in the original input)."""
        wb_bytes = _build_audit_workbook(
            [
                (1, "S1", "TopicA", "YES", "Good"),
                (2, "S2", "TopicB", "", ""),
            ],
            # TopicC appears in Findings but has no sentences — must still
            # appear in selected_categories.
            findings_topics=["TopicA", "TopicB", "TopicC"],
        )
        result = audit.detect_partial_audit(wb_bytes)
        assert result["is_partial"] is True
        assert result["selected_categories"] == ["TopicA", "TopicB", "TopicC"]

    def test_selected_categories_fallback_when_findings_empty(self):
        """If Findings has only a header row, fall back to category_stats keys
        (preserves legacy behavior so callers always have a non-empty list
        when is_partial is True)."""
        wb_bytes = _build_audit_workbook(
            [
                (1, "S1", "TopicA", "YES", "Good"),
                (2, "S2", "TopicB", "", ""),
            ],
            findings_topics=None,
        )
        result = audit.detect_partial_audit(wb_bytes)
        assert result["is_partial"] is True
        assert set(result["selected_categories"]) == {"TopicA", "TopicB"}

    def test_selected_categories_empty_when_not_audit_format(self):
        wb_bytes = _make_workbook_bytes({
            "Data": [["Col1", "Col2"], ["v1", "v2"]],
        })
        result = audit.detect_partial_audit(wb_bytes)
        assert result["selected_categories"] == []


# ===========================================================================
# _merge_settings_history
# ===========================================================================

class TestMergeSettingsHistory:
    def test_appends_new_value(self):
        assert audit._merge_settings_history("anthropic", "openai") == "anthropic; openai"

    def test_preserves_order(self):
        assert audit._merge_settings_history("a; b", "c") == "a; b; c"

    def test_dedupes_existing(self):
        assert audit._merge_settings_history("anthropic", "anthropic") == "anthropic"

    def test_dedupes_within_history(self):
        assert audit._merge_settings_history("a; b; a", "c") == "a; b; c"

    def test_empty_existing(self):
        assert audit._merge_settings_history("", "anthropic") == "anthropic"
        assert audit._merge_settings_history(None, "anthropic") == "anthropic"

    def test_empty_current(self):
        assert audit._merge_settings_history("anthropic", "") == "anthropic"
        assert audit._merge_settings_history("anthropic", None) == "anthropic"

    def test_strips_whitespace_in_existing(self):
        assert audit._merge_settings_history("  a ;  b  ", "c") == "a; b; c"

    def test_both_empty(self):
        assert audit._merge_settings_history("", "") == ""
        assert audit._merge_settings_history(None, None) == ""


# ===========================================================================
# _build_category_sentences
# ===========================================================================

class TestBuildCategorySentences:
    def test_input_format(self):
        df = pd.DataFrame(
            [[1, "Food is good", "Food"], [2, "Slow service", "Service"]],
            columns=["#", "Sentences", "Category"],
        )
        result = audit._build_category_sentences(df, is_output_format=False)
        assert "Food" in result
        assert "Service" in result
        assert result["Food"][1] == "Food is good"
        assert result["Service"][2] == "Slow service"

    def test_output_format_with_id(self):
        df = pd.DataFrame(
            [[1, "Food is good", "Food", "YES", "Correct"],
             [2, "Slow service", "Service", "NO", "Wrong"]],
            columns=["ID", "Sentence", "Topic", "Audit", "Explanation"],
        )
        result = audit._build_category_sentences(df, is_output_format=True)
        assert "Food" in result
        assert result["Food"][1] == "Food is good"

    def test_output_format_without_id(self):
        df = pd.DataFrame(
            [["Food is good", "Food", "YES", "Correct"],
             ["Slow service", "Service", "NO", "Wrong"]],
            columns=["Sentence", "Topic", "Audit", "Explanation"],
        )
        result = audit._build_category_sentences(df, is_output_format=True)
        assert "Food" in result
        assert "Service" in result
        # Uses synthetic row-based IDs
        assert len(result["Food"]) == 1
        assert len(result["Service"]) == 1

    def test_skips_nan_rows(self):
        df = pd.DataFrame(
            [[1, "Food is good", "Food"], [2, None, "Service"], [3, "Nice", None]],
            columns=["#", "Sentences", "Category"],
        )
        result = audit._build_category_sentences(df, is_output_format=False)
        # Only the first row has all required fields
        assert "Food" in result
        assert "Service" not in result
        # Row 3 has no category
        total_sentences = sum(len(v) for v in result.values())
        assert total_sentences == 1

    def test_blank_first_row_raises(self):
        df = pd.DataFrame(
            [[None, None, None], [1, "Food is good", "Food"]],
            columns=["#", "Sentences", "Category"],
        )
        with pytest.raises(ValueError, match="first row.*blank"):
            audit._build_category_sentences(df, is_output_format=False)


# ===========================================================================
# _apply_precision_formula
# ===========================================================================

class TestPrecisionFormula:
    def test_formula_structure(self):
        wb = Workbook()
        ws_sentences = wb.active
        ws_sentences.title = "Sentences"
        ws_sentences.append(["ID", "Sentence", "Topic", "Audit", "Explanation"])
        ws_sentences.append([1, "S1", "TopicA", "YES", "Good"])
        ws_sentences.append([2, "S2", "TopicA", "NO", "Bad"])
        ws_sentences.append([3, "S3", "TopicA", "", ""])

        ws_findings = wb.create_sheet("Findings")
        ws_findings.append(["Topic", "Description", "Accuracy", "Issues"])
        ws_findings.append(["TopicA", "Desc", "", ""])

        audit._apply_precision_formula(ws_findings, 2, "Sentences")

        formula = ws_findings.cell(row=2, column=3).value
        assert "YES" in formula
        assert formula.count("COUNTIFS") == 2
        wb.close()

    def test_formula_percentage_format(self):
        wb = Workbook()
        ws_findings = wb.active
        ws_findings.title = "Findings"
        ws_findings.append(["Topic", "Description", "Accuracy", "Issues"])
        ws_findings.append(["TopicA", "Desc", "", ""])
        wb.create_sheet("Sentences")

        audit._apply_precision_formula(ws_findings, 2, "Sentences")
        cell = ws_findings.cell(row=2, column=3)
        assert cell.number_format == numbers.FORMAT_PERCENTAGE
        wb.close()


# ===========================================================================
# _write_settings_sheet / _update_setting
# ===========================================================================

class TestSettingsSheet:
    def test_write_settings(self):
        wb = Workbook()
        ws = audit._ensure_settings_sheet(wb)
        audit._write_settings_sheet(ws, {"Model": "gpt-4o", "Provider": "openai"})
        assert ws.cell(row=2, column=1).value == "Model"
        assert ws.cell(row=2, column=2).value == "gpt-4o"
        wb.close()

    def test_write_replaces_existing(self):
        wb = Workbook()
        ws = audit._ensure_settings_sheet(wb)
        audit._write_settings_sheet(ws, {"Key1": "V1", "Key2": "V2"})
        audit._write_settings_sheet(ws, {"Key3": "V3"})
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "Key3"
        wb.close()

    def test_update_setting(self):
        wb = Workbook()
        ws = audit._ensure_settings_sheet(wb)
        audit._write_settings_sheet(ws, {"Model": "gpt-4o", "Provider": "openai"})
        audit._update_setting(ws, "Model", "claude-sonnet-4-20250514")
        assert ws.cell(row=2, column=2).value == "claude-sonnet-4-20250514"
        assert ws.cell(row=3, column=2).value == "openai"
        wb.close()

    def test_update_missing_key(self):
        wb = Workbook()
        ws = audit._ensure_settings_sheet(wb)
        audit._write_settings_sheet(ws, {"Model": "gpt-4o"})
        audit._update_setting(ws, "Nonexistent", "value")
        assert ws.cell(row=2, column=2).value == "gpt-4o"
        wb.close()


# ===========================================================================
# _write_errors_sheet
# ===========================================================================

class TestErrorsSheet:
    def test_creates_sheet(self):
        wb = Workbook()
        audit._write_errors_sheet(wb, ["Warning one", "Warning two"])
        assert "Errors" in wb.sheetnames
        ws = wb["Errors"]
        assert ws.cell(row=2, column=2).value == "Warning one"
        assert ws.cell(row=3, column=2).value == "Warning two"
        wb.close()

    def test_empty_list_no_sheet(self):
        wb = Workbook()
        audit._write_errors_sheet(wb, [])
        assert "Errors" not in wb.sheetnames
        wb.close()

    def test_removes_existing_on_empty(self):
        wb = Workbook()
        audit._write_errors_sheet(wb, ["Warning"])
        assert "Errors" in wb.sheetnames
        audit._write_errors_sheet(wb, [])
        assert "Errors" not in wb.sheetnames
        wb.close()


# ===========================================================================
# _format_categories_selected
# ===========================================================================

class TestFormatCategoriesSelected:
    def test_all_none(self):
        assert audit._format_categories_selected(None, 50, 50) == "All (50)"

    def test_all_empty(self):
        assert audit._format_categories_selected([], 50, 50) == "All (50)"

    def test_subset(self):
        assert audit._format_categories_selected(["A", "B"], 2, 100) == "2 of 100"
