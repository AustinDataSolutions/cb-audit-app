"""Tests for audit-report-summarizer.py — summary generation."""
from __future__ import annotations

import time
from io import BytesIO
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from tests.conftest import _load_module, _make_workbook_bytes

summarizer = _load_module("audit-report-summarizer.py", "audit_report_summarizer")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_completed_audit():
    """Build a completed audit workbook suitable for summarize_audit_report."""
    wb = Workbook()
    ws_sentences = wb.active
    ws_sentences.title = "Sentences"
    ws_sentences.append(["Sentence ID", "Sentence", "Topic", "Audit", "Explanation"])
    ws_sentences.append([1, "Sentence one", "Topic A", "YES", "Looks accurate"])
    ws_sentences.append([2, "Sentence two", "Topic A", "NO", "Wrong classification"])

    ws_topics = wb.create_sheet("Topics")
    ws_topics.append(["Topic", "Description", "Accuracy"])
    ws_topics.append(["Topic A", "Desc", 0.5])

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.getvalue()


# ===========================================================================
# _parse_llm_summary
# ===========================================================================

class TestParseLlmSummary:
    def test_valid_format(self):
        response = "Some preamble text.\nSUMMARY: The model struggles with topic X."
        result = summarizer._parse_llm_summary(response, lambda *a: None)
        assert result == "The model struggles with topic X."

    def test_case_insensitive(self):
        response = "summary: Lower case summary here."
        result = summarizer._parse_llm_summary(response, lambda *a: None)
        assert result == "Lower case summary here."

    def test_multiline_after_summary(self):
        response = "SUMMARY: Line one\nLine two\nLine three"
        result = summarizer._parse_llm_summary(response, lambda *a: None)
        assert "Line one" in result
        assert "Line two" in result

    def test_missing_summary_prefix(self):
        response = "No expected sections here."
        result = summarizer._parse_llm_summary(response, lambda *a: None)
        assert result == "REGEX FAILED TO PARSE LLM RESPONSE"

    def test_empty_input(self):
        result = summarizer._parse_llm_summary("", lambda *a: None)
        assert result == "REGEX FAILED TO PARSE LLM RESPONSE"

    def test_logs_warning_on_failure(self):
        log_calls = []
        summarizer._parse_llm_summary("no match", lambda msg: log_calls.append(msg))
        assert any("REGEX FAILED" in msg for msg in log_calls)


# ===========================================================================
# _call_llm_with_status (summarizer version)
# ===========================================================================

class TestCallLlmWithStatusSummarizer:
    def test_fast_call_no_status(self):
        status_calls = []
        result = summarizer._call_llm_with_status(
            lambda: "fast", 300, lambda msg: status_calls.append(msg)
        )
        assert result == "fast"
        assert status_calls == [""]

    def test_exception_reraised(self):
        def fail():
            raise ValueError("boom")

        with pytest.raises(ValueError, match="boom"):
            summarizer._call_llm_with_status(fail, 300, None)

    def test_slow_call_triggers_status(self):
        original = summarizer._LLM_STATUS_DELAY
        summarizer._LLM_STATUS_DELAY = 1
        status_calls = []

        def slow():
            time.sleep(2.5)
            return "result"

        try:
            result = summarizer._call_llm_with_status(
                slow, 300, lambda msg: status_calls.append(msg)
            )
            assert result == "result"
            waiting_msgs = [m for m in status_calls if m.startswith("Waiting")]
            assert len(waiting_msgs) >= 1
            assert status_calls[-1] == ""
        finally:
            summarizer._LLM_STATUS_DELAY = original


# ===========================================================================
# _coerce_audit_bytes
# ===========================================================================

class TestCoerceAuditBytes:
    def test_bytes(self):
        assert summarizer._coerce_audit_bytes(b"hello") == b"hello"

    def test_bytesio(self):
        assert summarizer._coerce_audit_bytes(BytesIO(b"data")) == b"data"

    def test_file_like(self, tmp_path):
        p = tmp_path / "f.bin"
        p.write_bytes(b"file-data")
        with p.open("rb") as fh:
            assert summarizer._coerce_audit_bytes(fh) == b"file-data"

    def test_path_string(self, tmp_path):
        p = tmp_path / "f.bin"
        p.write_bytes(b"path-data")
        assert summarizer._coerce_audit_bytes(str(p)) == b"path-data"

    def test_none_raises(self):
        with pytest.raises(ValueError):
            summarizer._coerce_audit_bytes(None)

    def test_invalid_type_raises(self):
        with pytest.raises(TypeError):
            summarizer._coerce_audit_bytes(12345)


# ===========================================================================
# _build_audit_findings
# ===========================================================================

class TestBuildAuditFindings:
    def test_basic(self):
        df = pd.DataFrame([
            [1, "S1", "Cat A", "NO", "Wrong"],
            [2, "S2", "Cat A", "YES", "Right"],
            [3, "S3", "Cat B", "YES", "Good"],
        ])
        findings = summarizer._build_audit_findings(df)
        assert "Cat A" in findings
        assert "Cat B" in findings
        assert findings["Cat A"][1] == ("NO", "Wrong")

    def test_skips_nan_rows(self):
        df = pd.DataFrame([
            [1, "S1", "Cat A", "NO", "Wrong"],
            [2, None, "Cat A", "YES", "Right"],
        ])
        findings = summarizer._build_audit_findings(df)
        assert 1 in findings["Cat A"]
        assert 2 not in findings.get("Cat A", {})


# ===========================================================================
# summarize_audit_report — integration-level test
# ===========================================================================

class TestSummarizeAuditReport:
    def test_adds_issues_column(self):
        wb_bytes = _build_completed_audit()
        updated = summarizer.summarize_audit_report(
            audit_excel_input=wb_bytes,
            msg_template="",
            llm_provider="anthropic",
            accuracy_threshold=0,  # skip LLM calls
        )
        wb = load_workbook(BytesIO(updated))
        ws = wb["Topics"]
        headers = [cell.value for cell in ws[1]]
        assert "Issues" in headers
        wb.close()
