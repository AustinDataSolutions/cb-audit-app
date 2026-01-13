from __future__ import annotations

import argparse
from io import BytesIO
import os
import re

import anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook
import pandas as pd
import yaml

try:
    from openai import OpenAI
except ImportError:  # Optional dependency
    OpenAI = None


DEFAULT_MODEL = "claude-sonnet-4-5"
DEFAULT_OPENAI_MODEL = "gpt-5-nano"
DEFAULT_MAX_TOKENS = 10000
DEFAULT_ACCURACY_THRESHOLD = 0.80


class SummaryStopRequested(Exception):
    """Raised when the user requests to stop the summary generation."""
    pass


def _load_prompts(prompts_path):
    with open(prompts_path, 'r') as f:
        return yaml.safe_load(f) or {}


def _load_config(config_path):
    with open(config_path, 'r') as f:
        return yaml.safe_load(f) or {}


def _get_llm_client(llm_provider, anthropic_api_key=None, openai_api_key=None):
    load_dotenv()

    if llm_provider == "anthropic":
        api_key = anthropic_api_key or os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise RuntimeError("ANTHROPIC_API_KEY environment variable not set")
        return anthropic.Anthropic(api_key=api_key)

    if llm_provider == "openai":
        if OpenAI is None:
            raise RuntimeError("openai package is not installed")
        api_key = openai_api_key or os.getenv('OPENAI_API_KEY')
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY environment variable not set")
        return OpenAI(api_key=api_key)

    raise ValueError("llm_provider must be 'anthropic' or 'openai'")


def _coerce_audit_bytes(audit_excel_input):
    if audit_excel_input is None:
        raise ValueError("audit_excel_input is required")

    if isinstance(audit_excel_input, (bytes, bytearray)):
        return bytes(audit_excel_input)

    if isinstance(audit_excel_input, BytesIO):
        return audit_excel_input.getvalue()

    if hasattr(audit_excel_input, "read"):
        return audit_excel_input.read()

    if isinstance(audit_excel_input, (str, os.PathLike)):
        with open(audit_excel_input, "rb") as f:
            return f.read()

    raise TypeError("audit_excel_input must be bytes, BytesIO, file-like, or a path")


def _build_audit_findings(df):
    if len(df.columns) < 4:
        raise ValueError("Audit report must contain at least four columns.")

    has_id_column = len(df.columns) >= 5
    if has_id_column:
        id_col = df.columns[0]
        sentence_col = df.columns[1]
        category_col = df.columns[2]
        judgment_col = df.columns[3]
        explanation_col = df.columns[4]
    else:
        id_col = None
        sentence_col = df.columns[0]
        category_col = df.columns[1]
        judgment_col = df.columns[2]
        explanation_col = df.columns[3]

    audit_findings = {}
    for idx, row in df.iterrows():
        category = row[category_col]
        sentence = row[sentence_col]
        judgment = row[judgment_col]
        explanation = row[explanation_col]
        if has_id_column:
            sentence_id = row[id_col]
        else:
            sentence_id = f"row-{idx + 2}"
        if (
            pd.isna(category)
            or pd.isna(sentence)
            or (has_id_column and pd.isna(sentence_id))
            or pd.isna(judgment)
            or pd.isna(explanation)
        ):
            continue
        if category not in audit_findings:
            audit_findings[category] = {}
        audit_findings[category][sentence_id] = (judgment, explanation)
    return audit_findings


def _normalize_topic(value):
    text = str(value).strip()
    return " ".join(text.split())


def _topic_key(value):
    text = _normalize_topic(value)
    parts = [part.strip() for part in re.split(r"\s*-->\s*", text) if part.strip()]
    normalized = "-->".join(parts) if parts else text
    return normalized.casefold()


def _get_sentences_sheet_name(sheet_names):
    for name in sheet_names:
        if name.casefold() == "sentences":
            return name
    if not sheet_names:
        raise ValueError("Audit file has no worksheets.")
    return sheet_names[0]


def _get_topics_sheet_name(sheet_names):
    for name in sheet_names:
        lowered = name.casefold()
        if lowered in ("findings", "topics"):
            return name
    for name in sheet_names:
        if name.casefold() == "categories":
            return name
    raise ValueError("Audit file does not include a Topics/Findings worksheet.")


def _collect_summary_records(
    audit_findings,
    llm_provider,
    model_name,
    max_tokens,
    accuracy_threshold,
    model_info,
    msg_template,
    anthropic_api_key,
    openai_api_key,
    progress_fn,
    warn_fn,
    check_stop_fn=None,
):
    summary_records = []
    issues_by_key = {}
    total_categories = len(audit_findings)
    categories_checked = 1
    client = None

    for category, findings in audit_findings.items():
        # Check if stop was requested
        if check_stop_fn and check_stop_fn():
            raise SummaryStopRequested("Summary generation stopped by user request")

        if progress_fn:
            progress_fn(categories_checked, total_categories, category)
        inaccurate_sent_explanations = ""
        sent_count = 0
        wrong_count = 0

        for _, (judgment, explanation) in findings.items():
            sent_count += 1
            if judgment == "NO":
                wrong_count += 1
                inaccurate_sent_explanations += f"{explanation}\n"

        accuracy = round(((sent_count - wrong_count) / sent_count), 2) if sent_count else 0
        summary_text = ""
        if accuracy < accuracy_threshold:
            if client is None:
                client = _get_llm_client(llm_provider, anthropic_api_key, openai_api_key)
            message_content = msg_template.format(
                category=category,
                inaccurate_sent_explanations=inaccurate_sent_explanations,
                model_info=model_info or "",
            )
            if llm_provider == "anthropic":
                message = client.messages.create(
                    model=model_name,
                    max_tokens=max_tokens,
                    messages=[
                        {"role": "user", "content": message_content}
                    ]
                )
                response_text = message.content[0].text
            elif llm_provider == "openai":
                response = client.chat.completions.create(
                    model=model_name,
                    max_completion_tokens=max_tokens,
                    messages=[
                        {"role": "user", "content": message_content}
                    ]
                )
                response_text = response.choices[0].message.content
            else:
                raise ValueError("llm_provider must be 'anthropic' or 'openai'")
            summary_text = _parse_llm_summary(response_text, warn_fn)

        category_key = _topic_key(category)
        summary_records.append(
            {
                "category": category,
                "category_key": category_key,
                "accuracy": accuracy,
                "issues": summary_text,
            }
        )
        issues_by_key[category_key] = summary_text
        categories_checked += 1

    return summary_records, issues_by_key


def _parse_llm_summary(response_text, log_fn):
    pattern = r"SUMMARY:\s*(.+)"
    matches = re.findall(pattern, response_text, re.IGNORECASE | re.DOTALL)

    if matches:
        summary = matches[0]
        return summary.strip()

    log_fn("WARNING: REGEX FAILED TO PARSE LLM RESPONSE")
    summary_match = re.search(r"SUMMARY:\s*(.+?)$", response_text, re.IGNORECASE | re.DOTALL)
    if summary_match:
        return summary_match.group(1).strip()

    return "REGEX FAILED TO PARSE LLM RESPONSE"


def summarize_audit_report(
    audit_excel_input,
    msg_template=None,
    prompts_path=None,
    output_path=None,
    anthropic_api_key=None,
    openai_api_key=None,
    llm_provider="anthropic",
    model_name=DEFAULT_MODEL,
    max_tokens=DEFAULT_MAX_TOKENS,
    accuracy_threshold=DEFAULT_ACCURACY_THRESHOLD,
    model_info="",
    log_fn=None,
    warn_fn=None,
    progress_fn=None,
    check_stop_fn=None,
):
    if log_fn is None:
        log_fn = lambda *_args, **_kwargs: None
    if warn_fn is None:
        warn_fn = log_fn

    script_dir = os.path.dirname(os.path.abspath(__file__))
    prompts_path = prompts_path or os.path.join(script_dir, "prompts.yaml")
    if msg_template is None:
        prompts = _load_prompts(prompts_path)
        msg_template = prompts.get("summary_prompt", "")

    audit_bytes = _coerce_audit_bytes(audit_excel_input)
    excel_file = pd.ExcelFile(BytesIO(audit_bytes))
    sentences_sheet = _get_sentences_sheet_name(excel_file.sheet_names)
    df = pd.read_excel(excel_file, sheet_name=sentences_sheet)
    audit_findings = _build_audit_findings(df)

    if model_name is None:
        model_name = DEFAULT_MODEL if llm_provider == "anthropic" else DEFAULT_OPENAI_MODEL

    summary_records, issues_by_key = _collect_summary_records(
        audit_findings=audit_findings,
        llm_provider=llm_provider,
        model_name=model_name,
        max_tokens=max_tokens,
        accuracy_threshold=accuracy_threshold,
        model_info=model_info,
        msg_template=msg_template,
        anthropic_api_key=anthropic_api_key,
        openai_api_key=openai_api_key,
        progress_fn=progress_fn,
        warn_fn=warn_fn,
        check_stop_fn=check_stop_fn,
    )

    wb = load_workbook(BytesIO(audit_bytes))
    topics_sheet_name = _get_topics_sheet_name(wb.sheetnames)
    ws = wb[topics_sheet_name]
    header_cells = list(ws[1])
    headers = []
    for cell in header_cells:
        if cell.value is None:
            headers.append("")
        elif isinstance(cell.value, str):
            headers.append(cell.value.strip().casefold())
        else:
            headers.append(str(cell.value).strip().casefold())
    try:
        topic_col = headers.index("topic") + 1
    except ValueError as exc:
        wb.close()
        raise ValueError("Topics worksheet must include a 'Topic' column.") from exc

    issues_col = None
    if "issues" in headers:
        issues_col = headers.index("issues") + 1
    else:
        issues_col = len(headers) + 1
        ws.cell(row=1, column=issues_col, value="Issues")

    unmatched_topics = []
    used_keys = set()
    for row_idx in range(2, ws.max_row + 1):
        topic_value = ws.cell(row=row_idx, column=topic_col).value
        if topic_value is None:
            continue
        topic_text = str(topic_value).strip()
        if not topic_text:
            continue
        if topic_text.casefold() == "model average":
            continue
        topic_key = _topic_key(topic_text)
        if topic_key in issues_by_key:
            ws.cell(row=row_idx, column=issues_col, value=issues_by_key[topic_key])
            used_keys.add(topic_key)
        else:
            unmatched_topics.append(topic_text)

    missing_topics = [
        category for category_key, category in (
            (_topic_key(name), name) for name in audit_findings.keys()
        )
        if category_key not in used_keys
    ]
    if warn_fn:
        if unmatched_topics:
            warn_fn(
                "Summary topics not found in Topics worksheet:\n"
                + "\n".join(f"- {topic}" for topic in sorted(set(unmatched_topics)))
            )
        if missing_topics:
            warn_fn(
                "Audit categories missing from Topics worksheet:\n"
                + "\n".join(f"- {topic}" for topic in sorted(set(missing_topics)))
            )

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    output_bytes = output.getvalue()

    if output_path:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(output_bytes)

    return output_bytes


def main():
    parser = argparse.ArgumentParser(description="Summarize a completed audit report.")
    parser.add_argument("--input", dest="input_path", help="Path to the completed audit .xlsx file")
    parser.add_argument("--output", dest="output_path", help="Path to write the updated audit .xlsx file")
    parser.add_argument("--prompts", dest="prompts_path", help="Path to prompts.yaml")
    parser.add_argument("--config", dest="config_path", help="Path to config.yaml")
    parser.add_argument("--api-key", dest="api_key", help="Anthropic API key override")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    prompts_path = args.prompts_path or os.path.join(script_dir, "prompts.yaml")
    config_path = args.config_path or os.path.join(script_dir, "config.yaml")
    config = _load_config(config_path)

    input_path = args.input_path
    if not input_path:
        audit_file_name = config.get("cli_summary", {}).get("audit_file")
        if not audit_file_name:
            raise ValueError("No input path provided and cli_summary.audit_file is missing in config.yaml")
        input_path = os.path.join(script_dir, "inputs", audit_file_name)

    output_path = args.output_path or input_path

    def _log(msg):
        print(msg)

    summarize_audit_report(
        audit_excel_input=input_path,
        msg_template=_load_prompts(prompts_path).get("summary_prompt", ""),
        prompts_path=prompts_path,
        output_path=output_path,
        anthropic_api_key=args.api_key,
        log_fn=_log,
    )

    print(f"Workbook saved to {output_path}")


if __name__ == "__main__":
    main()
